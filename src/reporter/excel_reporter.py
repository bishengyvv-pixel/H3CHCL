import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

from src.models.assessment_result import AssessmentResult

logger = logging.getLogger(__name__)

# ── 配色体系 ──────────────────────────────────────────────
NAVY = "1B3A5C"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F6FA"
BORDER_GRAY = "D0D5DD"

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color=WHITE, size=10)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=18, color=NAVY)
SUBTITLE_FONT = Font(name="微软雅黑", bold=True, size=12, color="4A5568")
BODY_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_GRAY),
    right=Side(style="thin", color=BORDER_GRAY),
    top=Side(style="thin", color=BORDER_GRAY),
    bottom=Side(style="thin", color=BORDER_GRAY),
)

# 风险等级色
HIGH_FILL = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
HIGH_FONT = Font(name="微软雅黑", size=10, bold=True, color="C53030")
MEDIUM_FILL = PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid")
MEDIUM_FONT = Font(name="微软雅黑", size=10, bold=True, color="975A16")
LOW_FILL = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")
LOW_FONT = Font(name="微软雅黑", size=10, bold=True, color="276749")

# 角色色
ROLE_COLORS = {
    "core": PatternFill(start_color="E9D8FD", end_color="E9D8FD", fill_type="solid"),
    "aggregation": PatternFill(start_color="BEE3F8", end_color="BEE3F8", fill_type="solid"),
    "access": PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid"),
    "router": PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid"),
}
ROLE_LABELS = {
    "core": "核心层", "aggregation": "汇聚层",
    "access": "接入层", "router": "路由器", "unknown": "未知",
}

# 得分色阶
def score_fill(score: int) -> PatternFill:
    if score < 60:
        return HIGH_FILL
    elif score < 85:
        return MEDIUM_FILL
    return LOW_FILL


class ExcelReporter:
    """生成美观的安全评估 Excel 报告：仪表盘 + 角色汇总 + 设备详情。"""

    def __init__(self, output_dir: str | Path = "output") -> None:
        self._output_dir = Path(output_dir)

    def generate(self, results: list[AssessmentResult], filename: str = "assessment_report.xlsx") -> Path:
        self._output_dir.mkdir(parents=True, exist_ok=True)
        wb = Workbook()

        self._build_dashboard(wb, results)
        if results:
            self._build_role_summary(wb, results)
        self._build_detail(wb, results)

        filepath = self._output_dir / filename
        wb.save(filepath)
        logger.info("Excel 报告已保存 → %s", filepath)
        return filepath

    # ═══════════════════════════════════════════════════════════
    #  仪表盘
    # ═══════════════════════════════════════════════════════════
    def _build_dashboard(self, wb: Workbook, results: list[AssessmentResult]) -> None:
        ws = wb.active
        ws.title = "安全仪表盘"
        ws.sheet_properties.tabColor = NAVY

        total = len(results)
        if total == 0:
            ws.merge_cells("A1:C1")
            ws["A1"] = "无评估数据"
            ws["A1"].font = TITLE_FONT
            return

        high_cnt = sum(1 for r in results if r.risk_level == "high")
        medium_cnt = sum(1 for r in results if r.risk_level == "medium")
        low_cnt = sum(1 for r in results if r.risk_level == "low")
        avg_score = round(sum(r.score for r in results) / total, 1)

        # 标题
        ws.merge_cells("A1:F1")
        ws["A1"] = "H3C 园区网安全评估报告"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 35

        # ── 摘要卡片区 KPI ──
        cards = [
            ("设备总数", f"{total} 台", NAVY),
            ("平均得分", f"{avg_score} / 100", "276749" if avg_score >= 85 else "975A16" if avg_score >= 60 else "C53030"),
            ("高风险", f"{high_cnt} 台", "C53030" if high_cnt else BORDER_GRAY),
            ("中风险", f"{medium_cnt} 台", "975A16" if medium_cnt else BORDER_GRAY),
            ("低风险", f"{low_cnt} 台", "276749" if low_cnt else BORDER_GRAY),
        ]
        for i, (label, value, color) in enumerate(cards):
            col = i * 1 + 1
            ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col)
            cell_label = ws.cell(row=3, column=col, value=label)
            cell_label.font = Font(name="微软雅黑", size=9, color="718096")
            cell_label.alignment = CENTER
            cell_label.border = THIN_BORDER
            ws.row_dimensions[3].height = 20

            ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col)
            cell_val = ws.cell(row=4, column=col, value=value)
            cell_val.font = Font(name="微软雅黑", bold=True, size=16, color=color)
            cell_val.alignment = CENTER
            cell_val.border = THIN_BORDER
            ws.row_dimensions[4].height = 36

        # 卡片区底部分隔
        ws.row_dimensions[5].height = 8

        # ── 风险分布饼图 ──
        ws.merge_cells("A7:C7")
        ws["A7"] = "风险分布"
        ws["A7"].font = SUBTITLE_FONT

        chart_data_start = 8
        ws.cell(row=chart_data_start, column=1, value="等级").font = BODY_FONT
        ws.cell(row=chart_data_start, column=2, value="数量").font = BODY_FONT
        for i, (label, cnt) in enumerate([("高风险", high_cnt), ("中风险", medium_cnt), ("低风险", low_cnt)], 1):
            ws.cell(row=chart_data_start + i, column=1, value=label).font = BODY_FONT
            ws.cell(row=chart_data_start + i, column=2, value=cnt).font = BODY_FONT

        pie = PieChart()
        pie.title = "风险分布"
        pie.width = 14
        pie.height = 10
        pie_data = Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_start + 3)
        pie_labels = Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_start + 3)
        pie.add_data(pie_data, titles_from_data=False)
        pie.set_categories(pie_labels)

        # 饼图配色
        colors = ["C53030", "D69E2E", "38A169"]
        for idx in range(3):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = colors[idx]
            pie.series[0].data_points.append(pt)

        ws.add_chart(pie, "D7")

        # ── 按角色汇总 ──
        role_offset_col = 1
        ws.merge_cells(start_row=7, start_column=role_offset_col, end_row=7, end_column=role_offset_col + 1)
        # Already has "风险分布" at A7

        # ── 角色分布表 ──
        role_row = chart_data_start + 5
        if not results:
            return
        ws.merge_cells(f"A{role_row}:C{role_row}")
        ws.cell(row=role_row, column=1, value="按角色分布").font = SUBTITLE_FONT

        role_dist: dict[str, int] = {}
        role_scores: dict[str, list[int]] = {}
        for r in results:
            label = ROLE_LABELS.get(r.role, r.role)
            role_dist[label] = role_dist.get(label, 0) + 1
            role_scores.setdefault(label, []).append(r.score)

        role_row += 1
        headers = ["角色", "数量", "平均分"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=role_row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        for role_label, cnt in role_dist.items():
            role_row += 1
            avg = round(sum(role_scores[role_label]) / cnt, 1) if cnt else 0
            ws.cell(row=role_row, column=1, value=role_label).font = BODY_FONT
            ws.cell(row=role_row, column=2, value=cnt).font = BODY_FONT
            score_cell = ws.cell(row=role_row, column=3, value=avg)
            score_cell.font = BOLD_FONT
            score_cell.fill = score_fill(int(avg))
            for ci in range(1, 4):
                ws.cell(row=role_row, column=ci).border = THIN_BORDER
                ws.cell(row=role_row, column=ci).alignment = CENTER

        # 列宽
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12

    # ═══════════════════════════════════════════════════════════
    #  角色汇总
    # ═══════════════════════════════════════════════════════════
    def _build_role_summary(self, wb: Workbook, results: list[AssessmentResult]) -> None:
        ws = wb.create_sheet("角色安全分析")
        ws.sheet_properties.tabColor = "6B46C1"

        # 按角色排序
        role_order = {"core": 0, "aggregation": 1, "access": 2, "router": 3}
        results_sorted = sorted(results, key=lambda r: role_order.get(r.role, 99))

        # 统计每个角色最常出现的扣分项
        from collections import Counter

        role_failures: dict[str, Counter] = {}
        for r in results_sorted:
            label = ROLE_LABELS.get(r.role, r.role)
            role_failures.setdefault(label, Counter())
            for item in r.failed_items:
                role_failures[label][item.desc] += 1

        ws.merge_cells("A1:E1")
        ws["A1"] = "角色安全分析"
        ws["A1"].font = TITLE_FONT
        ws.row_dimensions[1].height = 35

        headers = ["角色", "设备数", "平均分", "最高频扣分项", "出现次数"]
        row = 3
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 24

        role_dist: dict[str, list[AssessmentResult]] = {}
        for r in results_sorted:
            label = ROLE_LABELS.get(r.role, r.role)
            role_dist.setdefault(label, []).append(r)

        for role_label, devs in role_dist.items():
            row += 1
            avg_s = round(sum(d.score for d in devs) / len(devs), 1)
            top_fail = role_failures[role_label].most_common(1)
            top_desc = top_fail[0][0] if top_fail else "—"
            top_cnt = top_fail[0][1] if top_fail else 0

            row_data = [role_label, len(devs), avg_s, top_desc, top_cnt]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                if ci == 3:
                    cell.font = BOLD_FONT
                    cell.fill = score_fill(int(avg_s))
                cell.alignment = LEFT_WRAP if ci == 4 else CENTER

        # 角色平均分柱状图
        chart_row = row + 3
        if len(role_dist) > 0:
            chart = BarChart()
            chart.title = "各角色平均得分"
            chart.width = 18
            chart.height = 10
            chart.y_axis.title = "平均得分"
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 100
            data_ref = Reference(ws, min_col=3, min_row=3, max_row=row)
            cat_ref = Reference(ws, min_col=1, min_row=4, max_row=row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cat_ref)
            ws.add_chart(chart, f"A{chart_row}")

        # 列宽
        widths = [14, 10, 10, 32, 12]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    # ═══════════════════════════════════════════════════════════
    #  设备详情
    # ═══════════════════════════════════════════════════════════
    def _build_detail(self, wb: Workbook, results: list[AssessmentResult]) -> None:
        ws = wb.create_sheet("设备详情")
        ws.sheet_properties.tabColor = "3182CE"

        # 按风险等级、角色排序
        risk_order = {"high": 0, "medium": 1, "low": 2}
        role_order = {"core": 0, "aggregation": 1, "access": 2, "router": 3}
        results_sorted = sorted(results, key=lambda r: (risk_order.get(r.risk_level, 99), role_order.get(r.role, 99)))

        # 标题
        ws.merge_cells("A1:I1")
        ws["A1"] = "设备安全评估明细"
        ws["A1"].font = TITLE_FONT
        ws.row_dimensions[1].height = 35

        # 表头
        headers = ["设备IP", "主机名", "角色", "得分", "风险等级", "扣分项ID", "扣分描述", "权重", "加固建议"]
        header_row = 3
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        ws.row_dimensions[header_row].height = 24

        # 冻结表头
        ws.freeze_panes = f"A{header_row + 1}"

        # 数据
        row = header_row + 1
        striped = False
        for result in results_sorted:
            role_label = ROLE_LABELS.get(result.role, result.role)
            items = result.failed_items if result.failed_items else []
            role_fill = ROLE_COLORS.get(result.role, PatternFill())

            if not items:
                # 全部通过
                dev_data = [result.device_ip, result.hostname, role_label, result.score, result.risk_level,
                            "—", "全部通过 ✓", 0, "—"]
                for ci, val in enumerate(dev_data, 1):
                    cell = ws.cell(row=row, column=ci, value=val)
                    cell.font = BOLD_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER if ci != 9 else LEFT_WRAP
                    if ci == 4:
                        cell.fill = score_fill(result.score)
                    elif ci == 5:
                        self._apply_risk_style(cell, result.risk_level)
                row += 1
                striped = not striped
                continue

            first_item_row = row
            for idx, item in enumerate(items):
                if idx == 0:
                    # 首行含设备概要（合并列暂用重复值）
                    dev_data = [result.device_ip, result.hostname, role_label, result.score, result.risk_level,
                                item.rule_id, item.desc, item.weight, item.fix_template]
                    for ci, val in enumerate(dev_data, 1):
                        cell = ws.cell(row=row, column=ci, value=val)
                        cell.font = BODY_FONT
                        cell.border = THIN_BORDER
                        cell.alignment = CENTER if ci <= 8 else LEFT_WRAP
                        if ci == 3:
                            cell.fill = role_fill
                        elif ci == 4:
                            cell.font = BOLD_FONT
                            cell.fill = score_fill(result.score)
                        elif ci == 5:
                            self._apply_risk_style(cell, result.risk_level)
                else:
                    # 后续扣分项行
                    item_data = ["", "", "", "", "", item.rule_id, item.desc, item.weight, item.fix_template]
                    for ci, val in enumerate(item_data, 1):
                        cell = ws.cell(row=row, column=ci, value=val)
                        cell.font = BODY_FONT
                        cell.border = THIN_BORDER
                        cell.alignment = CENTER if ci <= 8 else LEFT_WRAP
                row += 1

            # 合并设备信息列
            if len(items) > 1:
                for merge_col in [1, 2, 3, 4, 5]:
                    ws.merge_cells(start_row=first_item_row, start_column=merge_col,
                                   end_row=row - 1, end_column=merge_col)

            striped = not striped

        # 自动列宽（按内容适配）
        col_widths = {
            1: 16, 2: 14, 3: 10, 4: 8, 5: 10,
            6: 14, 7: 28, 8: 8, 9: 48,
        }
        for ci, w in col_widths.items():
            ws.column_dimensions[get_column_letter(ci)].width = w

        # 行高
        for r in range(header_row + 1, row):
            ws.row_dimensions[r].height = 26

    @staticmethod
    def _apply_risk_style(cell, risk_level: str) -> None:
        if risk_level == "high":
            cell.value = "高风险"
            cell.fill = HIGH_FILL
            cell.font = HIGH_FONT
        elif risk_level == "medium":
            cell.value = "中风险"
            cell.fill = MEDIUM_FILL
            cell.font = MEDIUM_FONT
        else:
            cell.value = "低风险"
            cell.fill = LOW_FILL
            cell.font = LOW_FONT
